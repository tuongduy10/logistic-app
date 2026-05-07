"""
direct_inventory_export.py
--------------------------
Computes inventory output (EOQ, Safety Stock, ROP, Total Cost, CV, Variability,
Demand Level) DIRECTLY from input_3rd.xlsx — bypassing the forecast pipeline.

Logic mirror of inventory.py / pipeline.py but uses ACTUAL historical demand
from sales_history instead of forecast values.

Existing pipeline code is NOT modified.
"""
from __future__ import annotations

import math
import os
import sys
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Side, Border
)
from openpyxl.utils import get_column_letter
from config import (
    COLORS, FORECAST_HORIZON, INPUT_FILE,
    MPL_THEME, OUTPUT_CHARTS_DIR, OUTPUT_FILES_DIR,
)

# ── 1. Load data ──────────────────────────────────────────────────────────────

def load_data(path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    sales = pd.read_excel(path, sheet_name="sales_history")
    params = pd.read_excel(path, sheet_name="inventory_parameters")
    sales.columns  = [c.strip().lower().replace(" ", "_") for c in sales.columns]
    params.columns = [c.strip().lower().replace(" ", "_") for c in params.columns]
    return sales, params


# ── 2. Compute inventory metrics (mirrors inventory.py logic exactly) ─────────

def coefficient_of_variation(mean: float, std: float) -> float:
    return round(std / mean, 4) if mean > 0 else 0.0


def variability_label(cv: float) -> str:
    return "Stable" if cv < 0.3 else "Volatile"


def demand_level_label(annual: float, p33: float, p67: float) -> str:
    if annual >= p67:
        return "High"
    if annual >= p33:
        return "Medium"
    return "Low"


def compute_inventory_rows(sales: pd.DataFrame, params: pd.DataFrame) -> pd.DataFrame:
    # Aggregate actual demand per SKU
    demand_stats = (
        sales.groupby(["sku_id"])["actual_demand_units"]
        .agg(demand_mean="mean", demand_std="std", count="count")
        .reset_index()
    )
    demand_stats["demand_std"] = demand_stats["demand_std"].fillna(0.0)

    # Merge with inventory params
    df = demand_stats.merge(params, on="sku_id", how="left")

    rows = []
    for _, r in df.iterrows():
        demand_mean   = float(r["demand_mean"])
        demand_std    = float(r["demand_std"])
        annual_demand = round(demand_mean * 12, 2)
        lead_time_mo  = float(r["lead_time_days_mean"]) / 30.0
        order_cost    = float(r["ordering_cost_usd_per_order"])
        holding_cost  = float(r["holding_cost_usd_per_unit_year"])
        z             = float(r["z_value"])

        eoq           = math.sqrt((2 * annual_demand * order_cost) / holding_cost) if holding_cost > 0 else 0.0
        safety_stock  = z * demand_std * math.sqrt(lead_time_mo)
        rop           = demand_mean * lead_time_mo + safety_stock
        total_cost    = (annual_demand / eoq) * order_cost + (eoq / 2) * holding_cost if eoq > 0 else 0.0
        cv            = coefficient_of_variation(demand_mean, demand_std)

        rows.append({
            "sku_id":               r["sku_id"],
            "sku_name":             r.get("sku_name_x") or r.get("sku_name", ""),
            "data_source":          "actual_historical",
            "months_of_data":       int(r["count"]),
            "annual_demand":        round(annual_demand, 2),
            "demand_mean_monthly":  round(demand_mean, 2),
            "demand_std_monthly":   round(demand_std, 2),
            "cv":                   cv,
            "variability":          variability_label(cv),
            "safety_stock":         round(safety_stock, 2),
            "rop":                  round(rop, 2),
            "eoq":                  round(eoq, 2),
            "total_cost":           round(total_cost, 2),
            "demand_level":         "",   # filled after percentile calc
        })

    result = pd.DataFrame(rows)

    # Percentile-based demand level (same logic as pipeline.py Stage 2)
    annual = result["annual_demand"].values
    p33 = float(np.percentile(annual, 33))
    p67 = float(np.percentile(annual, 67))
    result["demand_level"] = result["annual_demand"].apply(
        lambda d: demand_level_label(d, p33, p67)
    )

    return result, p33, p67


# ── 3. Export to Excel with formatting ───────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")   # dark navy
SUBHDR_FILL   = PatternFill("solid", start_color="2E75B6")   # blue
ALT_FILL      = PatternFill("solid", start_color="EBF3FB")   # light blue
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
HIGH_FILL     = PatternFill("solid", start_color="C44E52")
MED_FILL      = PatternFill("solid", start_color="DD8452")
LOW_FILL      = PatternFill("solid", start_color="4C72B0")
STABLE_FILL   = PatternFill("solid", start_color="55A868")
VOLATILE_FILL = PatternFill("solid", start_color="C44E52")

WHITE_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BOLD_FONT     = Font(bold=True, name="Arial", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
TITLE_FONT    = Font(bold=True, name="Arial", size=13, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def _cell(ws, row, col, value, font=None, fill=None, alignment=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font = font
    if fill:          c.fill = fill
    if alignment:     c.alignment = alignment
    if number_format: c.number_format = number_format
    c.border = THIN_BORDER
    return c


def export_excel(result: pd.DataFrame, p33: float, p67: float, output_path: str) -> None:
    wb = Workbook()

    # ── Sheet 1: inventory_output ──────────────────────────────────────────
    ws = wb.active
    ws.title = "inventory_output"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16

    headers = [
        ("SKU ID",              12),
        ("SKU Name",            22),
        ("Data Source",         16),
        ("Months",               8),
        ("Annual Demand",       14),
        ("Mean Monthly",        13),
        ("Std Monthly",         12),
        ("CV",                   8),
        ("Variability",         12),
        ("Safety Stock",        13),
        ("ROP",                 10),
        ("EOQ",                 10),
        ("Total Cost (USD/yr)", 18),
        ("Demand Level",        13),
    ]

    HDR_ROW = 3
    for col_idx, (hdr, width) in enumerate(headers, start=1):
        _cell(ws, HDR_ROW, col_idx, hdr, font=WHITE_FONT, fill=HEADER_FILL, alignment=CENTER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[HDR_ROW].height = 20

    # Data rows
    for r_idx, row in enumerate(result.itertuples(index=False), start=HDR_ROW + 1):
        fill = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL

        def dc(col, val, fmt=None, custom_fill=None, custom_font=None):
            _cell(ws, r_idx, col, val,
                  font=custom_font or NORMAL_FONT,
                  fill=custom_fill or fill,
                  alignment=CENTER,
                  number_format=fmt)

        dc(1,  row.sku_id)
        dc(2,  row.sku_name, custom_fill=fill)
        _cell(ws, r_idx, 2, row.sku_name,
              font=NORMAL_FONT, fill=fill,
              alignment=Alignment(horizontal="left", vertical="center"),
              number_format=None)
        ws.cell(r_idx, 2).border = THIN_BORDER

        dc(3,  row.data_source)
        dc(4,  row.months_of_data)
        dc(5,  row.annual_demand,        fmt='#,##0.00')
        dc(6,  row.demand_mean_monthly,  fmt='#,##0.00')
        dc(7,  row.demand_std_monthly,   fmt='#,##0.00')
        dc(8,  row.cv,                   fmt='0.0000')
        dc(9,  row.safety_stock,         fmt='#,##0.00')
        dc(10, row.rop,                  fmt='#,##0.00')
        dc(11, row.eoq,                  fmt='#,##0.00')
        dc(12, row.total_cost,           fmt='"$"#,##0.00')

        # Variability badge
        v_fill = STABLE_FILL if row.variability == "Stable" else VOLATILE_FILL
        _cell(ws, r_idx, 9 - 1, row.variability,   # col 9 = variability
              font=WHITE_FONT, fill=v_fill, alignment=CENTER)
        # fix: re-map correct columns
        # col mapping: 1=sku_id,2=sku_name,3=data_source,4=months,5=annual,
        #              6=mean,7=std,8=cv,9=variability,10=safety_stock,
        #              11=rop,12=eoq,13=total_cost,14=demand_level
        # Reset 9 to variability (overwrite above dc(9))
        _cell(ws, r_idx, 9,  row.variability,
              font=WHITE_FONT, fill=v_fill, alignment=CENTER)
        # overwrite 10-12 for safety stock/rop/eoq with correct values
        dc(10, row.safety_stock, fmt='#,##0.00')
        dc(11, row.rop,          fmt='#,##0.00')
        dc(12, row.eoq,          fmt='#,##0.00')
        dc(13, row.total_cost,   fmt='"$"#,##0.00')

        # Demand level badge
        dl = row.demand_level
        dl_fill = {"High": HIGH_FILL, "Medium": MED_FILL, "Low": LOW_FILL}.get(dl, fill)
        _cell(ws, r_idx, 14, dl, font=WHITE_FONT, fill=dl_fill, alignment=CENTER)

        ws.row_dimensions[r_idx].height = 18

    # Freeze header
    ws.freeze_panes = "A4"

    # ── Sheet 2: summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    summary_rows = [
        ("Number of SKUs",              len(result)),
        ("Total Annual Demand",         f"{result['annual_demand'].sum():,.0f} units"),
        ("Total EOQ (sum)",             f"{result['eoq'].sum():,.0f} units"),
        ("Total Safety Stock (sum)",    f"{result['safety_stock'].sum():,.0f} units"),
        ("Total Inventory Cost (sum)",  f"${result['total_cost'].sum():,.0f}"),
        ("Avg CV across SKUs",          f"{result['cv'].mean():.4f}"),
        ("Stable SKUs",                 len(result[result['variability']=='Stable'])),
        ("Volatile SKUs",               len(result[result['variability']=='Volatile'])),
        ("High Demand SKUs",            len(result[result['demand_level']=='High'])),
        ("Medium Demand SKUs",          len(result[result['demand_level']=='Medium'])),
        ("Low Demand SKUs",             len(result[result['demand_level']=='Low'])),
        ("Demand p33 threshold",        f"{p33:,.0f}"),
        ("Demand p67 threshold",        f"{p67:,.0f}"),
        ("Calculation method",          "Actual historical demand (no forecast)"),
    ]

    for i, (label, value) in enumerate(summary_rows, start=2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        _cell(ws2, i, 1, label, font=BOLD_FONT,  fill=fill, alignment=LEFT)
        _cell(ws2, i, 2, value, font=NORMAL_FONT, fill=fill, alignment=CENTER)
        ws2.row_dimensions[i].height = 18

    wb.save(output_path)
    print(f"✓ Exported: {output_path}")


# ── 4. Main ───────────────────────────────────────────────────────────────────

def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir    = os.path.join(os.path.dirname(__file__), "output_files")
    os.makedirs(out_dir, exist_ok=True)
    output_path = os.path.join(out_dir, f"direct_inventory_{timestamp}.xlsx")

    print(f"[Direct Inventory] Loading: {input_path}")
    sales, params = load_data(input_path)
    print(f"  → {sales['sku_id'].nunique()} SKUs | {len(sales)} rows in sales_history")

    result, p33, p67 = compute_inventory_rows(sales, params)
    print(f"  → p33={p33:.0f}  p67={p67:.0f}")
    print(result[["sku_id","sku_name","annual_demand","cv","variability","demand_level",
                  "safety_stock","rop","eoq","total_cost"]].to_string(index=False))

    export_excel(result, p33, p67, output_path)
    return output_path


if __name__ == "__main__":
    main()